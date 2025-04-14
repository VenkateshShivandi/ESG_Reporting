"""
Excel Parser Module

This module provides functionality to parse Excel files using multiple libraries:
- pandas: Efficient data reading, returning a DataFrame
- openpyxl: Fine control for complex cases

The module implements a complete ETL (Extract, Transform, Load) pipeline for Excel data:

1. Extract: Reads data from Excel files using pandas or openpyxl
2. Transform: Cleans and transforms the data through multiple steps:
   - Standardizing column names
   - Handling missing values and duplicates
   - Removing NA fields
   - Handling outliers
   - Converting numerical data to appropriate types
   - Merging sheets with common schemas (when applicable)
3. Load: Returns the processed data in a structured format

The main function is parse_excel(), which orchestrates the entire ETL process
and returns the data in a structured format.
"""

import os
import logging
from typing import Dict, List, Any, Union, Tuple

# Import utility functions
from .utils import safe_parse, create_result_dict

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Try to import the required libraries
try:
    import pandas as pd
    import numpy as np
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    logger.warning("pandas not available. Excel parsing will be limited.")

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available. Complex Excel parsing will be limited.")

def extract_data_with_pandas(file_path: str) -> Dict[str, Any]:
    """
    Extract data from an Excel file using pandas.
    
    Args:
        file_path (str): Path to the Excel file
        
    Returns:
        Dict[str, Any]: Dictionary with sheet names as keys and DataFrames as values
    """
    if not PANDAS_AVAILABLE:
        return {}
    
    try:
        # Read all sheets
        excel_file = pd.ExcelFile(file_path)
        sheet_names = excel_file.sheet_names
        
        result = {}
        for sheet_name in sheet_names:
            # Read sheet into DataFrame
            df = excel_file.parse(sheet_name)
            
            # Convert DataFrame to dict
            sheet_data = {
                "headers": df.columns.tolist(),
                "data": df.values.tolist()
            }
            
            result[sheet_name] = sheet_data
            
        return result
    except Exception as e:
        logger.error(f"Error extracting data with pandas: {str(e)}")
        return {}

def extract_data_with_openpyxl(file_path: str) -> Dict[str, Any]:
    """
    Extract data from an Excel file using openpyxl.
    
    Args:
        file_path (str): Path to the Excel file
        
    Returns:
        Dict[str, Any]: Dictionary with sheet names as keys and sheet data as values
    """
    if not OPENPYXL_AVAILABLE:
        return {}
    
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet_names = workbook.sheetnames
        
        result = {}
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            
            # Get header row (assume first row is header)
            headers = []
            for cell in sheet[1]:
                headers.append(str(cell.value) if cell.value is not None else "")
            
            # Get data rows
            data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                data.append(list(row))
            
            result[sheet_name] = {
                "headers": headers,
                "data": data
            }
        
        return result
    except Exception as e:
        logger.error(f"Error extracting data with openpyxl: {str(e)}")
        return {}

def extract_metadata(file_path: str) -> Dict[str, Any]:
    """
    Extract metadata from an Excel file.
    
    Args:
        file_path (str): Path to the Excel file
        
    Returns:
        Dict[str, Any]: Metadata dictionary
    """
    metadata = {
        "filename": os.path.basename(file_path),
        "filesize": os.path.getsize(file_path)
    }
    
    # Try to get sheet names
    sheet_names = []
    
    if PANDAS_AVAILABLE:
        try:
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names
        except Exception:
            pass
    
    if not sheet_names and OPENPYXL_AVAILABLE:
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            sheet_names = workbook.sheetnames
        except Exception:
            pass
    
    if sheet_names:
        metadata["sheet_names"] = sheet_names
        metadata["sheet_count"] = len(sheet_names)
    
    return metadata

def transform_numerical_data(sheets_data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Transform numerical data in the extracted sheets to appropriate numeric types.
    
    Args:
        sheets_data (Dict[str, Any]): Dictionary with sheet data extracted from Excel
        
    Returns:
        Dict[str, Any]: Transformed sheet data with proper numeric types
    """
    if not PANDAS_AVAILABLE:
        logger.warning("Pandas not available. Skipping numerical data transformation.")
        return sheets_data
    
    transformed_sheets = {}
    
    for sheet_name, sheet_content in sheets_data.items():
        # Get headers from current sheet data before centralization
        headers = sheet_content["headers"]  # Direct access from sheet content
        data = sheet_content["data"]
        
        # Skip empty sheets
        if not data or not headers:
            transformed_sheets[sheet_name] = sheet_content
            continue
        
        # If a DataFrame is already available, use it
        if "dataframe" in sheet_content:
            df = sheet_content["dataframe"]
        else:
            # Validate and align headers and data
            max_data_columns = max([len(row) for row in data]) if data else 0
            header_count = len(headers)
            
            if header_count != max_data_columns:
                logger.warning(f"Column count mismatch in sheet '{sheet_name}': {header_count} headers vs {max_data_columns} data columns")
                
                if header_count > max_data_columns:
                    # Trim extra headers
                    logger.info(f"Trimming headers from {header_count} to {max_data_columns} columns")
                    headers = headers[:max_data_columns]
                else:
                    # Extend headers with placeholder names
                    logger.info(f"Extending headers from {header_count} to {max_data_columns} columns")
                    headers.extend([f"column_{i+1}" for i in range(header_count, max_data_columns)])
            
            # Ensure all data rows have the same number of columns
            normalized_data = []
            for row in data:
                if len(row) < len(headers):
                    # Pad missing values with None
                    normalized_data.append(row + [None] * (len(headers) - len(row)))
                else:
                    # Trim extra values
                    normalized_data.append(row[:len(headers)])
            
            # Convert to DataFrame for easier processing
            try:
                df = pd.DataFrame(normalized_data, columns=headers)
            except Exception as e:
                logger.error(f"Error creating DataFrame for sheet '{sheet_name}': {str(e)}")
                transformed_sheets[sheet_name] = sheet_content
                continue
        
        # Identify numerical columns
        numerical_columns = []
        for col in df.columns:
            # Check if column contains numerical data
            if df[col].dropna().apply(lambda x: isinstance(x, (int, float))).all():
                numerical_columns.append(col)
        
        # Transform numerical columns
        for col in numerical_columns:
            try:
                # Check if values are integers or floats
                if df[col].dropna().apply(lambda x: float(x) == int(float(x))).all():
                    # All values are integers
                    df[col] = pd.to_numeric(df[col], errors='coerce').astype('Int64')
                else:
                    # Some values are floats
                    df[col] = pd.to_numeric(df[col], errors='coerce').astype('float')
                
                logger.info(f"Transformed column '{col}' in sheet '{sheet_name}' to numeric type")
            except Exception as e:
                logger.warning(f"Failed to transform column '{col}' in sheet '{sheet_name}': {str(e)}")
        
        # Keep headers in transformed data for final aggregation
        transformed_sheets[sheet_name] = {
            "headers": headers,  # Maintain temporarily for final aggregation
            "data": df.astype(object).where(pd.notnull(df), None).values.tolist()
        }
        
        # Keep the dataframe if it was in the original content
        if "dataframe" in sheet_content:
            transformed_sheets[sheet_name]["dataframe"] = df
    
    return transformed_sheets

def clean_and_transform_data(sheets_data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Clean and transform the extracted data by handling missing values, duplicates, outliers,
    standardizing headers, and removing NA fields.
    
    This function performs the following transformations:
    1. Standardizes column names (lowercase, spaces to underscores)
    2. Handles missing values by dropping rows that are entirely empty
    3. Removes columns that are entirely NA
    4. Removes duplicate rows
    5. Handles outliers in numerical columns using IQR method
    6. Merges sheets with identical schemas (if applicable)
    
    Args:
        sheets_data (Dict[str, Any]): Dictionary with sheet data extracted from Excel
        
    Returns:
        Dict[str, Any]: Cleaned and transformed sheet data
    """
    if not PANDAS_AVAILABLE:
        logger.warning("Pandas not available. Skipping data cleaning and transformation.")
        return sheets_data
    
    cleaned_sheets = {}
    
    for sheet_name, sheet_content in sheets_data.items():
        headers = sheet_content["headers"]
        data = sheet_content["data"]
        
        # Skip empty sheets
        if not data or not headers:
            cleaned_sheets[sheet_name] = sheet_content
            continue
        
        # Validate and align headers and data
        max_data_columns = max([len(row) for row in data]) if data else 0
        header_count = len(headers)
        
        if header_count != max_data_columns:
            logger.warning(f"Column count mismatch in sheet '{sheet_name}': {header_count} headers vs {max_data_columns} data columns")
            
            if header_count > max_data_columns:
                # Trim extra headers
                logger.info(f"Trimming headers from {header_count} to {max_data_columns} columns")
                headers = headers[:max_data_columns]
            else:
                # Extend headers with placeholder names
                logger.info(f"Extending headers from {header_count} to {max_data_columns} columns")
                headers.extend([f"column_{i+1}" for i in range(header_count, max_data_columns)])
        
        # Ensure all data rows have the same number of columns
        normalized_data = []
        for row in data:
            if len(row) < len(headers):
                # Pad missing values with None
                normalized_data.append(row + [None] * (len(headers) - len(row)))
            else:
                # Trim extra values
                normalized_data.append(row[:len(headers)])
        
        # Convert to DataFrame for easier processing
        try:
            df = pd.DataFrame(normalized_data, columns=headers)
            
            # 1. Standardize column names
            # Convert to lowercase, replace spaces with underscores, remove special characters
            df.columns = [str(col).lower().strip()
                          .replace(' ', '_')
                          .replace('-', '_')
                          .translate({ord(c): None for c in '!@#$%^&*()[]{};:,./<>?\\|`~=+'})  # Remove all special chars
                          for col in df.columns]
            
            standardized_headers = df.columns.tolist()
            
            # 2. Handle missing values
            # Replace empty strings with NaN for consistent handling
            df = df.replace('', np.nan)
            
            # Drop rows where all values are NaN
            df = df.dropna(how='all')
            
            # 3. Remove columns that are entirely NA
            df = df.dropna(axis=1, how='all')
            
            # Update headers after column removal
            standardized_headers = df.columns.tolist()
            
            # 4. Handle duplicates
            # Remove duplicate rows
            initial_row_count = len(df)
            df = df.drop_duplicates()
            if len(df) < initial_row_count:
                logger.info(f"Removed {initial_row_count - len(df)} duplicate rows from sheet '{sheet_name}'")
            
            # 5. Handle outliers for numerical columns
            numerical_cols = df.select_dtypes(include=['number']).columns
            for col in numerical_cols:
                try:
                    # Check if we have enough non-null values to calculate statistics
                    if df[col].count() > 10:  # Arbitrary threshold, adjust as needed
                        # Calculate Q1, Q3 and IQR
                        Q1 = df[col].quantile(0.25)
                        Q3 = df[col].quantile(0.75)
                        IQR = Q3 - Q1
                        
                        # Define bounds for outliers
                        lower_bound = Q1 - 1.5 * IQR
                        upper_bound = Q3 + 1.5 * IQR
                        
                        # Count outliers
                        outliers_count = df[(df[col] < lower_bound) | (df[col] > upper_bound)][col].count()
                        
                        # Replace outliers with NaN
                        if outliers_count > 0:
                            df.loc[(df[col] < lower_bound) | (df[col] > upper_bound), col] = np.nan
                            logger.info(f"Handled {outliers_count} outliers in column '{col}' in sheet '{sheet_name}'")
                except Exception as e:
                    logger.warning(f"Failed to handle outliers in column '{col}' in sheet '{sheet_name}': {str(e)}")
            
            # Store the cleaned data
            cleaned_sheets[sheet_name] = {
                "headers": standardized_headers,
                "data": df.values.tolist(),
                "dataframe": df  # Store DataFrame for potential merging or further processing
            }
        except Exception as e:
            logger.error(f"Error processing sheet '{sheet_name}': {str(e)}")
            # Store the original data with normalized columns to prevent further errors
            cleaned_sheets[sheet_name] = {
                "headers": headers,
                "data": normalized_data
            }
    
    # Check if sheets can be merged (have the same schema)
    if len(cleaned_sheets) > 1:
        sheet_schemas = {sheet_name: set(content["headers"]) for sheet_name, content in cleaned_sheets.items()}
        
        # Group sheets by schema
        schema_groups = {}
        for sheet_name, schema in sheet_schemas.items():
            schema_key = frozenset(schema)
            if schema_key not in schema_groups:
                schema_groups[schema_key] = []
            schema_groups[schema_key].append(sheet_name)
        
        # Merge sheets with the same schema
        for schema, sheet_group in schema_groups.items():
            if len(sheet_group) > 1:
                logger.info(f"Merging sheets with common schema: {', '.join(sheet_group)}")
                
                # Create a new merged sheet
                merged_df = pd.concat([cleaned_sheets[sheet]["dataframe"] for sheet in sheet_group])
                
                # Add a source column to track which sheet the data came from
                merged_df['source_sheet'] = [sheet for sheet in sheet_group for _ in range(len(cleaned_sheets[sheet]["dataframe"]))]
                
                # Remove duplicate rows after merging
                merged_df = merged_df.drop_duplicates()
                
                # Create a new sheet entry for the merged data
                merged_sheet_name = f"merged_{'_'.join(sheet_group)}"
                cleaned_sheets[merged_sheet_name] = {
                    "headers": merged_df.columns.tolist(),
                    "data": merged_df.values.tolist(),
                    "dataframe": merged_df,
                    "source_sheets": sheet_group
                }
                
                logger.info(f"Created merged sheet '{merged_sheet_name}' with {len(merged_df)} rows")
    
    # Remove the DataFrame objects before returning
    for sheet_name in cleaned_sheets:
        if "dataframe" in cleaned_sheets[sheet_name]:
            del cleaned_sheets[sheet_name]["dataframe"]
        if "source_sheets" in cleaned_sheets[sheet_name]:
            # Keep this information as metadata
            pass
    
    return cleaned_sheets

def _parse_excel_internal(file_path: str) -> Dict[str, Any]:
    """
    Internal function to parse an Excel file.
    
    Args:
        file_path (str): Path to the Excel file
        
    Returns:
        Dict[str, Any]: Parsed content
    """
    # Extract metadata
    metadata = extract_metadata(file_path)
    
    # Try pandas first
    sheets_data = {}
    if PANDAS_AVAILABLE:
        try:
            sheets_data = extract_data_with_pandas(file_path)
        except Exception as e:
            logger.warning(f"Failed to parse Excel with pandas: {str(e)}")
    
    # If pandas failed or is not available, try openpyxl
    if not sheets_data and OPENPYXL_AVAILABLE:
        try:
            sheets_data = extract_data_with_openpyxl(file_path)
        except Exception as e:
            logger.warning(f"Failed to parse Excel with openpyxl: {str(e)}")
    
    # If both methods failed
    if not sheets_data:
        return create_result_dict(error="Failed to parse Excel file with available libraries.")
    
    # Clean and transform the data
    cleaned_sheets = clean_and_transform_data(sheets_data)
    
    # Transform numerical data
    transformed_sheets = transform_numerical_data(cleaned_sheets)
    
    # Final aggregation with centralized headers
    return {
        "metadata": {
            **metadata,
            "columns": {sheet: content["headers"] for sheet, content in transformed_sheets.items()}
        },
        "data": {sheet: content["data"] for sheet, content in transformed_sheets.items()}
    }

def parse_excel(file_path: str) -> Dict[str, Any]:
    """
    Parse an Excel file and extract data and metadata.
    
    This function orchestrates the complete ETL process:
    1. Extract: Reads raw data from Excel files
    2. Transform: Cleans and transforms the data through multiple steps
       - Standardizing column names
       - Handling missing values and duplicates
       - Removing NA fields
       - Handling outliers
       - Converting numerical data to appropriate types
       - Merging sheets with common schemas (when applicable)
    3. Load: Returns the processed data in a structured format
    
    Args:
        file_path (str): Path to the Excel file
        
    Returns:
        Dict[str, Any]: A dictionary containing:
            - metadata (Dict[str, Any]): File metadata
            - sheets (Dict[str, Dict]): Cleaned and transformed sheet data
            - error (str): Error message (if any)
    """
    result = safe_parse(_parse_excel_internal, file_path)
    
    return result

# Example usage
if __name__ == "__main__":
    import sys
        
    file_path = sys.argv[1]
    result = parse_excel(file_path)