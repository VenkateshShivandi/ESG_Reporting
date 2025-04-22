"""
Excel Parser Module

This module provides functionality to parse Excel files using multiple libraries:
- pandas: Efficient data reading, returning a DataFrame
- openpyxl: Fine control for complex cases

The module implements a complete ETL (Extract, Transform, Load) pipeline for Excel data:

1. Extract: Reads data from Excel files using pandas or openpyxl
2. Transform: Cleans and transforms the data through multiple steps:
   - Standardizing column names
   - Handling missing values and duplicates
   - Removing NA fields
   - Handling outliers
   - Converting numerical data to appropriate types
   - Merging sheets with common schemas (when applicable)
3. Load: Returns the processed data in a structured format

The main function is parse_excel(), which orchestrates the entire ETL process
and returns the data in a structured format.
"""

import os
import logging
from typing import Dict, List, Any, Union, Tuple

# Import utility functions (Adjusted for new location)
from .utils import safe_parse, create_result_dict

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Try to import the required libraries
try:
    import pandas as pd
    import numpy as np
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    logger.warning("pandas not available. Excel parsing will be limited.")

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available. Complex Excel parsing will be limited.")

def extract_data_with_pandas(file_path: str) -> Dict[str, Any]:
    """
    Extract data from an Excel file using pandas.
    
    Args:
        file_path (str): Path to the Excel file
        
    Returns:
        Dict[str, Any]: Dictionary with sheet names as keys and sheet data (headers, data) as values
    """
    if not PANDAS_AVAILABLE:
        return {}
    
    try:
        # Read all sheets
        excel_file = pd.ExcelFile(file_path)
        sheet_names = excel_file.sheet_names
        
        result = {}
        for sheet_name in sheet_names:
            # Read sheet into DataFrame
            df = excel_file.parse(sheet_name)
            
            # Convert DataFrame to dict
            sheet_data = {
                "headers": df.columns.tolist(),
                "data": df.values.tolist()
            }
            
            result[sheet_name] = sheet_data
            
        return result
    except Exception as e:
        logger.error(f"Error extracting data with pandas: {str(e)}")
        return {}

def extract_data_with_openpyxl(file_path: str) -> Dict[str, Any]:
    """
    Extract data from an Excel file using openpyxl.
    
    Args:
        file_path (str): Path to the Excel file
        
    Returns:
        Dict[str, Any]: Dictionary with sheet names as keys and sheet data as values
    """
    if not OPENPYXL_AVAILABLE:
        return {}
    
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet_names = workbook.sheetnames
        
        result = {}
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            
            # Get header row (assume first row is header)
            headers = []
            for cell in sheet[1]:
                headers.append(str(cell.value) if cell.value is not None else "")
            
            # Get data rows
            data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                data.append(list(row))
            
            result[sheet_name] = {
                "headers": headers,
                "data": data
            }
        
        return result
    except Exception as e:
        logger.error(f"Error extracting data with openpyxl: {str(e)}")
        return {}

def extract_metadata(file_path: str) -> Dict[str, Any]:
    """
    Extract metadata from an Excel file.
    
    Args:
        file_path (str): Path to the Excel file
        
    Returns:
        Dict[str, Any]: Metadata dictionary
    """
    metadata = {
        "filename": os.path.basename(file_path),
        "filesize": os.path.getsize(file_path)
    }
    
    # Try to get sheet names
    sheet_names = []
    
    if PANDAS_AVAILABLE:
        try:
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names
        except Exception:
            pass
    
    if not sheet_names and OPENPYXL_AVAILABLE:
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            sheet_names = workbook.sheetnames
        except Exception:
            pass
    
    if sheet_names:
        metadata["sheet_names"] = sheet_names
        metadata["sheet_count"] = len(sheet_names)
    
    return metadata

def transform_numerical_data(sheets_data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Transform numerical data in the extracted sheets to appropriate numeric types.
    
    Args:
        sheets_data (Dict[str, Any]): Dictionary with sheet data extracted from Excel
        
    Returns:
        Dict[str, Any]: Transformed sheet data with proper numeric types
    """
    if not PANDAS_AVAILABLE:
        logger.warning("Pandas not available. Skipping numerical data transformation.")
        return sheets_data
    
    transformed_sheets = {}
    
    for sheet_name, sheet_content in sheets_data.items():
        # Get headers from current sheet data before centralization
        headers = sheet_content.get("headers", []) # Use .get for safety
        data = sheet_content.get("data", [])      # Use .get for safety
        
        # Skip empty sheets
        if not data or not headers:
            transformed_sheets[sheet_name] = sheet_content
            continue
        
        # If a DataFrame is already available, use it
        if "dataframe" in sheet_content:
            df = sheet_content["dataframe"]
        else:
            # Validate and align headers and data
            max_data_columns = max((len(row) for row in data), default=0)
            header_count = len(headers)
            
            if header_count != max_data_columns:
                logger.warning(f"Column count mismatch in sheet '{sheet_name}': {header_count} headers vs {max_data_columns} data columns")
                
                if header_count > max_data_columns:
                    # Trim extra headers
                    logger.info(f"Trimming headers from {header_count} to {max_data_columns} columns")
                    headers = headers[:max_data_columns]
                else:
                    # Extend headers with placeholder names
                    logger.info(f"Extending headers from {header_count} to {max_data_columns} columns")
                    headers.extend([f"column_{i+1}" for i in range(header_count, max_data_columns)])
            
            # Ensure all data rows have the same number of columns
            normalized_data = []
            for row in data:
                row_len = len(row)
                if row_len < len(headers):
                    # Pad missing values with None
                    normalized_data.append(list(row) + [None] * (len(headers) - row_len))
                elif row_len > len(headers):
                     # Trim extra values
                     normalized_data.append(list(row[:len(headers)]))
                else:
                    normalized_data.append(list(row))

            # Convert to DataFrame for easier processing
            try:
                df = pd.DataFrame(normalized_data, columns=headers)
            except Exception as e:
                logger.error(f"Error creating DataFrame for sheet '{sheet_name}': {str(e)}")
                transformed_sheets[sheet_name] = sheet_content
                continue
        
        # Identify numerical columns
        numerical_columns = []
        for col in df.columns:
             # Check if column might be numerical (ignoring NaNs)
             try:
                 pd.to_numeric(df[col], errors='raise')
                 numerical_columns.append(col)
             except (ValueError, TypeError):
                 pass # Column is not purely numerical

        # Transform numerical columns
        for col in numerical_columns:
            try:
                # Convert to numeric, coercing errors to NaN
                numeric_col = pd.to_numeric(df[col], errors='coerce')
                
                # Check if all non-NaN values are integers
                is_integer = numeric_col.dropna().apply(lambda x: float(x).is_integer()).all()

                if is_integer:
                    # All values are integers (or NaN)
                    df[col] = numeric_col.astype('Int64') # Use pandas nullable integer type
                else:
                    # Some values are floats (or NaN)
                    df[col] = numeric_col.astype('float')
                
                logger.info(f"Transformed column '{col}' in sheet '{sheet_name}' to numeric type: {df[col].dtype}")
            except Exception as e:
                logger.warning(f"Failed to transform column '{col}' in sheet '{sheet_name}': {str(e)}")
        
        # Store transformed data (including the DataFrame for potential later use)
        transformed_sheets[sheet_name] = {
            "headers": df.columns.tolist(),
            "data": df.values.tolist(),
            "dataframe": df # Store DataFrame for potentially merging sheets later
        }

    return transformed_sheets


def clean_and_transform_data(sheets_data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Clean and transform the extracted data by handling missing values, duplicates, outliers,
    standardizing headers, and removing NA fields.
    
    This function performs the following transformations:
    1. Standardizes column names (lowercase, spaces to underscores)
    2. Handles missing values by dropping rows that are entirely empty
    3. Removes columns that are entirely NA
    4. Removes duplicate rows
    5. Handles outliers in numerical columns using IQR method
    6. Merges sheets with identical schemas (if applicable)
    
    Args:
        sheets_data (Dict[str, Any]): Dictionary with sheet data extracted from Excel
        
    Returns:
        Dict[str, Any]: Cleaned and transformed sheet data
    """
    if not PANDAS_AVAILABLE:
        logger.warning("Pandas not available. Skipping data cleaning and transformation.")
        return sheets_data
    
    cleaned_sheets = {}
    
    for sheet_name, sheet_content in sheets_data.items():
        headers = sheet_content.get("headers", [])
        data = sheet_content.get("data", [])
        
        if not data or not headers:
            cleaned_sheets[sheet_name] = sheet_content
            continue
            
        try:
            # Convert to DataFrame for easier processing
            # Ensure all data rows have the same number of columns as headers
            normalized_data = []
            header_count = len(headers)
            for row in data:
                 row_len = len(row)
                 if row_len < header_count:
                     normalized_data.append(list(row) + [None] * (header_count - row_len))
                 elif row_len > header_count:
                     normalized_data.append(list(row[:header_count]))
                 else:
                     normalized_data.append(list(row))

            df = pd.DataFrame(normalized_data, columns=headers)

            # 1. Standardize column names
            df.columns = [str(col).lower().strip()
                          .replace(' ', '_')
                          .replace('-', '_')
                          .translate({ord(c): None for c in '!@#$%^&*()[]{};:,./<>?\\|`~=+'}) 
                          for col in df.columns]
            standardized_columns = df.columns.tolist()

            # 2. Handle missing values
            df = df.replace('', np.nan).dropna(how='all')
            
            # 3. Remove columns that are entirely NA
            df = df.dropna(axis=1, how='all')
            standardized_columns = df.columns.tolist() # Update columns after potential removal

            # 4. Handle duplicates
            initial_row_count = len(df)
            df = df.drop_duplicates()
            if len(df) < initial_row_count:
                logger.info(f"Removed {initial_row_count - len(df)} duplicate rows from sheet '{sheet_name}'")

            # 5. Handle outliers for numerical columns
            numerical_cols = df.select_dtypes(include=np.number).columns
            for col in numerical_cols:
                if df[col].count() > 10: # Ensure enough data points
                    Q1 = df[col].quantile(0.25)
                    Q3 = df[col].quantile(0.75)
                    IQR = Q3 - Q1
                    lower_bound = Q1 - 1.5 * IQR
                    upper_bound = Q3 + 1.5 * IQR
                    outliers = df[(df[col] < lower_bound) | (df[col] > upper_bound)]
                    if not outliers.empty:
                        df.loc[outliers.index, col] = np.nan # Replace outliers with NaN
                        logger.info(f"Handled {len(outliers)} outliers in column '{col}' of sheet '{sheet_name}'")
            
            cleaned_sheets[sheet_name] = {
                "headers": standardized_columns,
                "data": df.values.tolist(),
                "dataframe": df # Keep DataFrame for potential merging
            }
            
        except Exception as e:
            logger.error(f"Error cleaning sheet '{sheet_name}': {str(e)}")
            cleaned_sheets[sheet_name] = sheet_content # Keep original if cleaning fails

    # 6. Merge sheets with identical schemas (Optional step, can be complex)
    # This step is simplified here. A robust implementation would compare columns and types carefully.
    merged_sheets = {}
    schema_map = {} # Map schema tuple to list of sheet names

    for sheet_name, content in cleaned_sheets.items():
        if "dataframe" in content:
            schema = tuple(sorted(content["headers"]))
            if schema not in schema_map:
                schema_map[schema] = []
            schema_map[schema].append(sheet_name)

    for schema, sheet_names in schema_map.items():
        if len(sheet_names) > 1:
            # Merge DataFrames
            merged_df = pd.concat([cleaned_sheets[name]["dataframe"] for name in sheet_names], ignore_index=True)
            merged_name = "_".join(sheet_names) # Create a merged name
            merged_sheets[merged_name] = {
                 "headers": merged_df.columns.tolist(),
                 "data": merged_df.values.tolist(),
                 "merged_from": sheet_names # Add info about source sheets
            }
            logger.info(f"Merged sheets with identical schema {schema}: {', '.join(sheet_names)} into {merged_name}")
            # Remove original sheets that were merged
            for name in sheet_names:
                del cleaned_sheets[name]
        elif "dataframe" in cleaned_sheets[sheet_names[0]]: # Handle single sheets with DataFrames
             # Convert single sheets back from DataFrame if they weren't merged
             single_sheet_name = sheet_names[0]
             df = cleaned_sheets[single_sheet_name]["dataframe"]
             merged_sheets[single_sheet_name] = {
                 "headers": df.columns.tolist(),
                 "data": df.values.tolist()
             }
             del cleaned_sheets[single_sheet_name] # Remove the entry with the DataFrame


    # Combine non-merged sheets (those without DataFrames or unique schemas) and merged sheets
    final_sheets = {**cleaned_sheets, **merged_sheets}
    
    return final_sheets


def _parse_excel_internal(file_path: str) -> Dict[str, Any]:
    """
    Internal function to parse an Excel file.
    
    Args:
        file_path (str): Path to the Excel file
        
    Returns:
        Dict[str, Any]: Parsed content including metadata and data per sheet
    """
    # Extract metadata
    metadata = extract_metadata(file_path)
    
    # Try pandas first
    sheets_data = {}
    if PANDAS_AVAILABLE:
        try:
            sheets_data = extract_data_with_pandas(file_path)
            logger.info(f"Successfully extracted data using pandas for {metadata.get('filename')}")
        except Exception as e:
            logger.warning(f"Failed to parse Excel with pandas: {str(e)}")
    
    # If pandas failed or is not available, try openpyxl
    if not sheets_data and OPENPYXL_AVAILABLE:
        try:
            sheets_data = extract_data_with_openpyxl(file_path)
            logger.info(f"Successfully extracted data using openpyxl for {metadata.get('filename')}")
        except Exception as e:
            logger.warning(f"Failed to parse Excel with openpyxl: {str(e)}")
    
    # If both methods failed
    if not sheets_data:
        return create_result_dict(error="Failed to parse Excel file with available libraries.")
    
    # Clean and transform the data
    cleaned_sheets = clean_and_transform_data(sheets_data)
    
    # Transform numerical data
    transformed_sheets = transform_numerical_data(cleaned_sheets)
    
    # Final aggregation
    # Return data in the desired format: metadata and data (dict of sheets)
    final_data = {}
    final_metadata = metadata.copy()
    final_metadata["columns"] = {} # Store columns per sheet in metadata

    for sheet_name, content in transformed_sheets.items():
        if "headers" in content and "data" in content:
             final_data[sheet_name] = content["data"] # Only store data rows
             final_metadata["columns"][sheet_name] = content["headers"] # Store headers in metadata

    return {
        "metadata": final_metadata,
        "data": final_data # Main data is dict of sheet_name: list_of_data_rows
    }

def parse_excel(file_path: str) -> Dict[str, Any]:
    """
    Parse an Excel file and extract data and metadata.
    
    This function orchestrates the complete ETL process:
    1. Extract: Reads raw data from Excel files
    2. Transform: Cleans and transforms the data through multiple steps
       - Standardizing column names
       - Handling missing values and duplicates
       - Removing NA fields
       - Handling outliers
       - Converting numerical data to appropriate types
       - Merging sheets with common schemas (when applicable)
    3. Load: Returns the processed data in a structured format
    
    Args:
        file_path (str): Path to the Excel file
        
    Returns:
        Dict[str, Any]: A dictionary containing:
            - metadata (Dict[str, Any]): File metadata (including sheet names and columns per sheet)
            - data (Dict[str, List[List]]): Dictionary where keys are sheet names (or merged sheet names) 
                                            and values are lists of data rows for that sheet.
            - error (str): Error message (if any)
    """
    # Use safe_parse from utils to handle file existence, permissions, etc.
    result = safe_parse(_parse_excel_internal, file_path)
    
    return result

# Example usage
if __name__ == "__main__":
    import sys
    import json # Import json for printing
        
    if len(sys.argv) < 2:
        print("Usage: python excel_parser.py <excel_file_path>")
        sys.exit(1)
        
    file_path = sys.argv[1]
    result = parse_excel(file_path)

    if "error" in result:
         print(f"Error: {result['error']}")
    else:
         # Pretty print the result
         print(json.dumps(result, indent=2, ensure_ascii=False)) 